"""
Main execution script for momentum factor strategy.

Orchestrates data loading, factor calculation, preprocessing,
backtesting, and output generation.

Usage:
    cd /Users/hedi/report_reproduce
    python3 -m src.momentum_factor.main
"""

from __future__ import annotations

import sys
import time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

# Ensure project root is on path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from common.backtest import performance_analysis
from common.utils import standardize_factor
from src.momentum_factor.config import CONFIG
from src.momentum_factor.strategy import (
    apply_universe_filter,
    build_monthly_panel,
    calculate_momentum_factors,
    calculate_purified_momentum,
    calculate_residual_momentum,
    calculate_trend_momentum_factors,
    construct_fama_french_factors,
    load_stock_info,
    load_valuation,
    prepare_daily_panel,
)
from common.data_loader import (
    get_month_end_trading_days,
    load_industry,
    load_st_data,
    load_suspend,
)

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=pd.errors.PerformanceWarning)


# =====================================================================
# Helper: prepare factor panel for backtest
# =====================================================================

def _prepare_factor_panel(
    factor_df: pd.DataFrame,
    monthly_panel: pd.DataFrame,
    factor_col: str,
    industry_df: pd.DataFrame,
    start_date: str,
    end_date: str,
) -> pd.DataFrame:
    """Merge factor values with forward returns, market cap, industry,
    apply standardization, and return the panel ready for backtesting.

    Args:
        factor_df: DataFrame with [stock_code, date, <factor_col>].
        monthly_panel: Monthly panel with monthly_return and market cap.
        factor_col: Name of the factor column.
        industry_df: Industry classification.
        start_date: Backtest start date.
        end_date: Backtest end date.

    Returns:
        Factor panel ready for performance_analysis().
    """
    # Compute forward return (next month's return)
    fwd_return = (
        monthly_panel[["stock_code", "date", "monthly_return"]]
        .sort_values(["stock_code", "date"])
        .assign(forward_return=lambda df: df.groupby("stock_code")["monthly_return"].shift(-1))
        [["stock_code", "date", "forward_return"]]
    )

    # Merge factor, forward return, market cap, industry
    panel = (
        factor_df[["stock_code", "date", factor_col]]
        .merge(fwd_return, on=["stock_code", "date"], how="inner")
        .merge(
            monthly_panel[["stock_code", "date", "negotiable_market_value"]],
            on=["stock_code", "date"],
            how="inner",
        )
        .merge(industry_df, on="stock_code", how="left")
    )

    # Filter to backtest period
    panel = panel[
        (panel["date"] >= pd.Timestamp(start_date))
        & (panel["date"] <= pd.Timestamp(end_date))
    ].copy()

    # Drop rows missing factor or forward return
    panel = panel.dropna(subset=[factor_col, "forward_return"])

    # Standardize factor: winsorize -> neutralize (market cap + industry) -> z-score
    panel = standardize_factor(
        panel,
        factor_col=factor_col,
        date_col="date",
        market_cap_col="negotiable_market_value",
        industry_col="first_industry_name",
        winsorize_method=CONFIG.winsorize_method,
    )

    # Rename standardized factor to 'factor' for backtest module
    panel = panel.rename(columns={"factor_std": "factor"})

    return panel


# =====================================================================
# Main
# =====================================================================

def main() -> None:
    """Execute full momentum factor backtest pipeline."""
    t0 = time.time()
    output_dir = CONFIG.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("  Momentum Factor Strategy -- Backtest Pipeline")
    print("=" * 70)

    # ------------------------------------------------------------------
    # Step 1: Load data
    # ------------------------------------------------------------------
    print("\n[Step 1] Loading data...")

    daily_panel = prepare_daily_panel(
        CONFIG.buffer_start, CONFIG.end_date, CONFIG.data_dir
    )
    print(f"  Daily panel: {daily_panel.shape}")

    # Extended month-end dates (including lookback period)
    month_end_all = get_month_end_trading_days(
        CONFIG.buffer_start, CONFIG.end_date, CONFIG.data_dir
    )
    print(f"  Month-end dates (with buffer): {len(month_end_all)}")

    # Backtest-period month-end dates
    month_end_bt = month_end_all[month_end_all >= pd.Timestamp(CONFIG.start_date)]
    print(f"  Month-end dates (backtest): {len(month_end_bt)}")

    # Build monthly panel
    monthly_panel = build_monthly_panel(daily_panel, month_end_all)
    print(f"  Monthly panel: {monthly_panel.shape}")

    # Load auxiliary data
    st_data = load_st_data(CONFIG.data_dir)
    suspend_data = load_suspend(CONFIG.data_dir)
    stock_info = load_stock_info(CONFIG.data_dir)
    industry_df = load_industry(CONFIG.data_dir, CONFIG.industry_standard)
    valuation = load_valuation(CONFIG.data_dir, CONFIG.buffer_start, CONFIG.end_date)
    print(f"  Industry: {industry_df.shape}, Valuation: {valuation.shape}")

    # Apply universe filter
    monthly_filtered = apply_universe_filter(
        monthly_panel, st_data, suspend_data, stock_info, CONFIG.min_list_days
    )
    print(f"  Monthly panel (filtered): {monthly_filtered.shape}")

    # ------------------------------------------------------------------
    # Step 2: Calculate factors
    # ------------------------------------------------------------------
    print("\n[Step 2] Calculating factors...")

    # 2.1 Raw momentum factors
    print("  2.1 Raw momentum factors...")
    raw_mom = calculate_momentum_factors(
        monthly_filtered, daily_panel, month_end_all, CONFIG.momentum_months
    )
    print(f"      Shape: {raw_mom.shape}")

    # 2.2 Trend momentum factors
    print("  2.2 Trend momentum factors...")
    trend_mom = calculate_trend_momentum_factors(
        daily_panel, month_end_all, CONFIG.ma_windows
    )
    print(f"      Shape: {trend_mom.shape}")

    # 2.3 Purified momentum
    print("  2.3 Purified momentum factor...")
    purified_mom = calculate_purified_momentum(
        monthly_filtered, daily_panel, month_end_all
    )
    print(f"      Shape: {purified_mom.shape}")

    # 2.4 Residual momentum (Fama-French)
    print("  2.4 Residual momentum factor (FF3)...")
    ff_factors = construct_fama_french_factors(
        monthly_filtered, valuation, month_end_all
    )
    print(f"      FF factors: {ff_factors.shape}")
    resid_mom = calculate_residual_momentum(
        monthly_filtered, ff_factors, CONFIG.residual_lookback_months
    )
    print(f"      Residual momentum: {resid_mom.shape}")

    # ------------------------------------------------------------------
    # Step 3: Backtest each factor
    # ------------------------------------------------------------------
    print("\n[Step 3] Running backtests...")

    # Define all factors to test
    factor_configs: list[tuple[str, str, pd.DataFrame]] = []

    # Raw momentum factors
    for n in CONFIG.momentum_months:
        col_name = f"Momentum_{n}M"
        factor_configs.append((col_name, col_name, raw_mom))

    # Momentum_1M_Max
    factor_configs.append(("Momentum_1M_Max", "Momentum_1M_Max", raw_mom))

    # Trend momentum
    for w in CONFIG.ma_windows:
        col_name = f"MA_{w}"
        factor_configs.append((col_name, col_name, trend_mom))

    # Purified momentum
    factor_configs.append(("Momentum_1M_Neu", "Momentum_1M_Neu", purified_mom))

    # Residual momentum
    factor_configs.append(("Momentum_1M_Resid", "Momentum_1M_Resid", resid_mom))

    all_results: dict[str, dict] = {}
    summary_records: list[dict] = []

    for factor_name, factor_col, factor_source in factor_configs:
        print(f"\n  >>> Backtesting: {factor_name}")

        # Prepare factor panel
        try:
            panel = _prepare_factor_panel(
                factor_source,
                monthly_filtered,
                factor_col,
                industry_df,
                CONFIG.start_date,
                CONFIG.end_date,
            )
        except Exception as e:
            print(f"      SKIP (prep error): {e}")
            continue

        if panel.empty or panel["factor"].dropna().shape[0] < 100:
            print(f"      SKIP (insufficient data): {panel.shape}")
            continue

        print(f"      Panel shape: {panel.shape}, "
              f"dates: {panel['date'].nunique()}, "
              f"stocks/date: {panel.groupby('date').size().mean():.0f}")

        # Run performance analysis
        factor_output_dir = output_dir / factor_name
        factor_output_dir.mkdir(parents=True, exist_ok=True)

        results = performance_analysis(
            panel,
            factor_col="factor",
            return_col="forward_return",
            date_col="date",
            n_groups=CONFIG.n_groups,
            transaction_cost=CONFIG.transaction_cost,
            output_dir=factor_output_dir,
            report_name=factor_name,
            periods_per_year=CONFIG.periods_per_year,
        )

        all_results[factor_name] = results

        # Collect summary
        ic_stats = results["ic_stats"]
        ls_perf = results["ls_result"]["ls_performance"]
        group_perf = results["group_perf_df"]

        # Compute Mono_Score: (G10_ann_ret - G1_ann_ret) / (G8_ann_ret - G3_ann_ret)
        try:
            g10_ret = group_perf.loc[f"G{CONFIG.n_groups}", "ann_return"]
            g1_ret = group_perf.loc["G1", "ann_return"]
            g8_ret = group_perf.loc["G8", "ann_return"]
            g3_ret = group_perf.loc["G3", "ann_return"]
            mono_score = (g10_ret - g1_ret) / (g8_ret - g3_ret) if abs(g8_ret - g3_ret) > 1e-8 else np.nan
        except KeyError:
            mono_score = np.nan

        # Average turnover (top group)
        turnover_data = results["ls_result"]["quantile_result"]["turnover"]
        top_group_key = f"G{CONFIG.n_groups}"
        avg_turnover = (
            np.mean(turnover_data[top_group_key])
            if top_group_key in turnover_data and turnover_data[top_group_key]
            else np.nan
        )

        summary_records.append({
            "Factor": factor_name,
            "IC": ic_stats["ic_mean"],
            "IC_IR": ic_stats["icir"],
            "LongShort_Sharpe": ls_perf["sharpe"],
            "Mono_Score": mono_score,
            "Turnover": avg_turnover,
            "AnnReturn_LS": ls_perf["ann_return"],
            "MaxDD_LS": ls_perf["max_drawdown"],
            "IC_WinRate": ic_stats["win_rate"],
        })

        print(f"      IC={ic_stats['ic_mean']:.4f}, "
              f"ICIR={ic_stats['icir']:.2f}, "
              f"LS_Sharpe={ls_perf['sharpe']:.2f}, "
              f"Mono={mono_score:.2f}")

    # ------------------------------------------------------------------
    # Step 4: Summary output
    # ------------------------------------------------------------------
    print("\n[Step 4] Generating summary output...")

    summary_df = pd.DataFrame(summary_records).set_index("Factor")
    print("\n" + "=" * 70)
    print("  Factor Comparison Summary")
    print("=" * 70)
    print(summary_df.to_string(float_format="%.4f"))

    # Save comprehensive Excel
    _save_comprehensive_excel(summary_df, all_results, output_dir)

    elapsed = time.time() - t0
    print(f"\n[Done] Total time: {elapsed:.1f}s")
    print(f"[Done] Results saved to: {output_dir}")


# =====================================================================
# Comprehensive Excel output
# =====================================================================

def _save_comprehensive_excel(
    summary_df: pd.DataFrame,
    all_results: dict[str, dict],
    output_dir: Path,
) -> None:
    """Save all results into a single comprehensive Excel workbook."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    excel_path = output_dir / "backtest_summary.xlsx"

    # --- Generate comparison charts ---
    _generate_comparison_charts(summary_df, all_results, output_dir)

    # --- Write data sheets ---
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Sheet 1: Factor comparison summary
        fmt_df = summary_df.copy()
        fmt_df.to_excel(writer, sheet_name="因子对比汇总")

        # Sheet 2: IC time series (all factors)
        ic_all = pd.DataFrame()
        for name, res in all_results.items():
            ic_s = res["ic_df"].set_index("date")["ic"].rename(name)
            ic_all = pd.concat([ic_all, ic_s], axis=1)
        ic_all.index.name = "日期"
        ic_all.to_excel(writer, sheet_name="IC序列")

        # Sheet 3: Group performance per factor
        row_offset = 0
        for name, res in all_results.items():
            gp = res["group_perf_df"].copy()
            gp.index.name = "分组"
            header_df = pd.DataFrame({"": [f"--- {name} ---"]})
            header_df.to_excel(writer, sheet_name="分组表现", startrow=row_offset, index=False)
            gp.to_excel(writer, sheet_name="分组表现", startrow=row_offset + 1)
            row_offset += len(gp) + 4

    # --- Embed images ---
    wb = load_workbook(excel_path)

    def _embed(ws, img_name: str, anchor: str) -> None:
        p = output_dir / img_name
        if p.exists():
            img = XlImage(str(p))
            img.width = 720
            img.height = 400
            ws.add_image(img, anchor)

    # Summary sheet: embed comparison charts
    ws_summary = wb["因子对比汇总"]
    _embed(ws_summary, "ic_comparison_bar.png", "H2")
    _embed(ws_summary, "mono_score_bar.png", "H24")

    # IC sheet: embed IC series overlay
    ws_ic = wb["IC序列"]
    _embed(ws_ic, "ic_series_comparison.png", "P2")

    # Format all sheets
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = bold_font
        # Auto-width
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)
        ws.freeze_panes = "A2"

    wb.save(excel_path)
    print(f"  Comprehensive Excel saved: {excel_path}")


def _generate_comparison_charts(
    summary_df: pd.DataFrame,
    all_results: dict[str, dict],
    output_dir: Path,
) -> None:
    """Generate cross-factor comparison charts."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import seaborn as sns

    sns.set_style("whitegrid")
    plt.rcParams["font.sans-serif"] = [
        "Songti SC", "Heiti TC", "STHeiti", "Kaiti SC", "PingFang HK", "SimHei"
    ]
    plt.rcParams["axes.unicode_minus"] = False

    BLUE = "#1f77b4"
    RED = "#d62728"

    # 1. IC comparison bar chart
    fig, ax = plt.subplots(figsize=(12, 6))
    factors = summary_df.index.tolist()
    ic_vals = summary_df["IC"].values * 100
    colors = [BLUE if v >= 0 else RED for v in ic_vals]
    bars = ax.bar(range(len(factors)), ic_vals, color=colors, alpha=0.8)
    ax.set_xticks(range(len(factors)))
    ax.set_xticklabels(factors, rotation=45, ha="right", fontsize=9)
    ax.set_ylabel("IC (%)", fontsize=12)
    ax.set_title("各因子IC均值对比", fontsize=14)
    ax.axhline(0, color="black", linewidth=0.5)

    # Add IC_IR as text on bars
    for i, (bar, icir) in enumerate(zip(bars, summary_df["IC_IR"].values)):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.1 * np.sign(bar.get_height()),
            f"ICIR={icir:.2f}",
            ha="center", va="bottom" if bar.get_height() >= 0 else "top",
            fontsize=8,
        )
    fig.tight_layout()
    fig.savefig(output_dir / "ic_comparison_bar.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 2. Mono score bar chart
    fig, ax = plt.subplots(figsize=(12, 6))
    mono_vals = summary_df["Mono_Score"].values
    colors = [BLUE if abs(v) >= 2 else RED for v in mono_vals]
    ax.bar(range(len(factors)), mono_vals, color=colors, alpha=0.8)
    ax.set_xticks(range(len(factors)))
    ax.set_xticklabels(factors, rotation=45, ha="right", fontsize=9)
    ax.set_ylabel("Mono Score", fontsize=12)
    ax.set_title("各因子单调性得分对比", fontsize=14)
    ax.axhline(1, color="gray", linestyle="--", linewidth=0.8, label="Mono=1 (基准)")
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(output_dir / "mono_score_bar.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 3. IC time-series comparison (overlay key factors)
    fig, ax = plt.subplots(figsize=(14, 6))
    key_factors = ["Momentum_1M", "MA_60", "Momentum_1M_Neu", "Momentum_1M_Resid"]
    line_styles = ["-", "--", "-.", ":"]
    line_colors = [RED, BLUE, "green", "purple"]

    for factor_name, ls, lc in zip(key_factors, line_styles, line_colors):
        if factor_name in all_results:
            ic_s = all_results[factor_name]["ic_df"].set_index("date")["ic"]
            cum_ic = ic_s.cumsum()
            ax.plot(cum_ic.index, cum_ic.values, label=factor_name,
                    linestyle=ls, color=lc, linewidth=1.5)

    ax.set_title("关键因子累计RankIC对比", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("累计RankIC", fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(output_dir / "ic_series_comparison.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


# =====================================================================
# Entry point
# =====================================================================

if __name__ == "__main__":
    main()
