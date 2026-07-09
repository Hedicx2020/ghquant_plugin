"""Run the report reproduction and generate checkable outputs.

Usage:
    python3 -m src.test.main
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.test.config import CONFIG
from src.test.strategy import BacktestConfig, build_backtest

BLUE = "#1f77b4"
RED = "#d62728"
GRAY = "#7f7f7f"

plt.style.use("seaborn-v0_8-whitegrid")
sns.set_palette("RdBu_r")
plt.rcParams["font.sans-serif"] = [
    "Songti SC",
    "Heiti TC",
    "STHeiti",
    "Kaiti SC",
    "PingFang HK",
    "SimHei",
    "Arial Unicode MS",
]
plt.rcParams["axes.unicode_minus"] = False


REPORT_BENCHMARK = pd.DataFrame(
    {
        "策略": ["周内+复合跷跷板+隔日反转(多空)", "周内+复合跷跷板+隔日反转(仅做多)", "T"],
        "cumulative_return": [0.8522, 0.5158, 0.2078],
        "annual_return": [0.0765, 0.0510, 0.0228],
        "max_drawdown": [0.0326, 0.0372, 0.0746],
        "annual_volatility": [0.0375, 0.0301, 0.0389],
        "calmar": [2.35, 1.37, 0.31],
        "sharpe": [2.04, 1.70, 0.59],
        "win_rate": [0.5631, np.nan, np.nan],
        "long_win_rate": [0.5729, np.nan, np.nan],
        "short_win_rate": [0.5690, np.nan, np.nan],
        "up_market_win_rate": [0.6289, np.nan, np.nan],
        "down_market_win_rate": [0.4902, np.nan, np.nan],
        "profit_loss_ratio": [1.44, 1.49, 1.11],
        "annual_timing_count": [129.48, np.nan, np.nan],
    }
).set_index("策略")


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 36)
    workbook.save(path)


def _embed_images(path: Path, image_map: dict[str, list[Path]]) -> None:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage

    workbook = load_workbook(path)
    for sheet_name, images in image_map.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        row = max(sheet.max_row + 3, 5)
        for image_path in images:
            if image_path.exists():
                img = XlImage(str(image_path))
                img.width = 720
                img.height = 360
                sheet.add_image(img, f"A{row}")
                row += 22
    workbook.save(path)


def _save_charts(results: dict[str, Any], output_dir: Path) -> None:
    backtest = results["backtest"]
    ic_series = results["ic_series"].set_index("date")
    group_perf = results["group_performance"]

    nav = (1.0 + backtest[["strategy_return", "long_only_return", "benchmark_return"]]).cumprod()
    nav.columns = ["多空策略", "仅做多策略", "T基准"]

    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(nav.index, nav["多空策略"], label="多空策略", color=BLUE, linewidth=1.8)
    ax.plot(nav.index, nav["仅做多策略"], label="仅做多策略", color="black", linewidth=1.4)
    ax.plot(nav.index, nav["T基准"], label="T基准", color=GRAY, linewidth=1.2)
    ax.set_title("净值对比", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("累计净值", fontsize=12)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(output_dir / "net_value_comparison.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(12, 6))
    corr = ic_series["rolling_signal_return_corr"]
    colors = [BLUE if value >= 0 else RED for value in corr.fillna(0)]
    ax.bar(corr.index, corr.values, color=colors, alpha=0.75, width=2.5)
    ax.axhline(0, color="black", linewidth=0.6)
    ax.set_title("信号与次日国债期货收益滚动相关", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("60日滚动相关系数", fontsize=12)
    fig.tight_layout()
    fig.savefig(output_dir / "ic_series.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    active_returns = backtest.loc[backtest["position"].ne(0), "strategy_return"]
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.hist(active_returns, bins=40, color=BLUE, alpha=0.75, edgecolor="white")
    ax.axvline(active_returns.mean(), color=RED, linestyle="--", linewidth=1.5, label="均值")
    ax.set_title("策略持仓日收益分布", fontsize=14)
    ax.set_xlabel("日收益率", fontsize=12)
    ax.set_ylabel("频数", fontsize=12)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(output_dir / "ic_distribution.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    source_returns = pd.DataFrame(index=backtest.index)
    for source in ["daily_upper_calendar", "reverse", "seesaw"]:
        source_returns[source] = backtest["strategy_return"].where(backtest["signal_source"].eq(source), 0.0)
    source_nav = (1.0 + source_returns).cumprod()
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = {"daily_upper_calendar": BLUE, "reverse": RED, "seesaw": "black"}
    labels = {"daily_upper_calendar": "日度上阈值+周内", "reverse": "隔日反转", "seesaw": "复合跷跷板"}
    for source, label in labels.items():
        ax.plot(source_nav.index, source_nav[source], label=label, color=colors[source], linewidth=1.5)
    ax.set_title("按信号来源累计收益", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("累计净值", fontsize=12)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(output_dir / "group_cumulative_returns.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 6))
    bar_data = group_perf.loc[["daily_upper_calendar", "reverse", "seesaw"], "annual_return"] * 100
    bar_colors = [BLUE if value >= 0 else RED for value in bar_data]
    ax.bar(["日度上阈值+周内", "隔日反转", "复合跷跷板"], bar_data.values, color=bar_colors, alpha=0.8)
    ax.set_title("按信号来源年化收益", fontsize=14)
    ax.set_xlabel("信号来源", fontsize=12)
    ax.set_ylabel("年化收益(%)", fontsize=12)
    fig.tight_layout()
    fig.savefig(output_dir / "group_returns_bar.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


def _comparison(summary: pd.DataFrame) -> pd.DataFrame:
    common_index = summary.index.intersection(REPORT_BENCHMARK.index)
    common_cols = summary.columns.intersection(REPORT_BENCHMARK.columns)
    reproduced = summary.loc[common_index, common_cols]
    benchmark = REPORT_BENCHMARK.loc[common_index, common_cols]
    diff = reproduced - benchmark
    denom = benchmark.abs().replace(0, np.nan)
    deviation = diff.abs() / denom

    rows = []
    for strategy in common_index:
        for metric in common_cols:
            rows.append(
                {
                    "策略": strategy,
                    "指标": metric,
                    "复现值": reproduced.loc[strategy, metric],
                    "报告值": benchmark.loc[strategy, metric],
                    "绝对偏差": diff.loc[strategy, metric],
                    "相对偏差": deviation.loc[strategy, metric],
                }
            )
    return pd.DataFrame(rows)


def _save_excel_outputs(results: dict[str, Any], output_dir: Path) -> None:
    summary = results["summary"]
    comparison = _comparison(summary)
    backtest = results["backtest"].copy()
    ic_series = results["ic_series"]
    group_perf = results["group_performance"]
    annual = results["annual_performance"]

    signal_stats = pd.DataFrame(
        {
            "指标": [
                "样本开始",
                "样本结束",
                "样本天数",
                "多头天数",
                "空头天数",
                "空仓天数",
                "日度上阈值+周内触发",
                "隔日反转触发",
                "复合跷跷板触发",
            ],
            "值": [
                str(backtest.index.min().date()),
                str(backtest.index.max().date()),
                len(backtest),
                int(backtest["position"].gt(0).sum()),
                int(backtest["position"].lt(0).sum()),
                int(backtest["position"].eq(0).sum()),
                int(backtest["signal_source"].eq("daily_upper_calendar").sum()),
                int(backtest["signal_source"].eq("reverse").sum()),
                int(backtest["signal_source"].eq("seesaw").sum()),
            ],
        }
    )

    outputs = {
        "backtest_summary.xlsx": {
            "summary": summary.reset_index(),
            "report_benchmark": REPORT_BENCHMARK.reset_index(),
            "comparison": comparison,
            "annual": annual.reset_index(names="year"),
        },
        "ic_series.xlsx": {
            "rolling_corr": ic_series,
            "returns": backtest.reset_index(names="date")[
                ["date", "position", "future_return", "strategy_return"]
            ],
        },
        "group_performance.xlsx": {
            "group_performance": group_perf.reset_index(),
            "source_daily_returns": backtest.reset_index(names="date")[
                ["date", "signal_source", "position", "strategy_return", "benchmark_return"]
            ],
        },
        "factor_statistics.xlsx": {
            "signal_stats": signal_stats,
            "signals": backtest.reset_index(names="date")[
                [
                    "date",
                    "daily_upper",
                    "daily_lower",
                    "ls_signal",
                    "reverse_signal",
                    "calendar_signal",
                    "daily_upper_calendar",
                    "seesaw_signal",
                    "position",
                    "signal_source",
                ]
            ],
        },
    }

    for filename, sheets in outputs.items():
        path = output_dir / filename
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        _format_workbook(path)

    consolidated = output_dir / "test_results.xlsx"
    with pd.ExcelWriter(consolidated, engine="openpyxl") as writer:
        summary.reset_index().to_excel(writer, sheet_name="summary", index=False)
        comparison.to_excel(writer, sheet_name="comparison", index=False)
        backtest.reset_index(names="date").to_excel(writer, sheet_name="net_value", index=False)
        ic_series.to_excel(writer, sheet_name="ic_series", index=False)
        group_perf.reset_index().to_excel(writer, sheet_name="group_performance", index=False)
        signal_stats.to_excel(writer, sheet_name="factor_statistics", index=False)
    _format_workbook(consolidated)
    _embed_images(
        consolidated,
        {
            "net_value": [output_dir / "net_value_comparison.png"],
            "ic_series": [output_dir / "ic_series.png", output_dir / "ic_distribution.png"],
            "group_performance": [
                output_dir / "group_cumulative_returns.png",
                output_dir / "group_returns_bar.png",
            ],
        },
    )


def _write_json_and_log(results: dict[str, Any], output_dir: Path) -> None:
    summary = results["summary"]
    comparison = _comparison(summary)
    metrics = {
        "config": results["config"].to_jsonable(),
        "summary": json.loads(summary.reset_index().to_json(orient="records", force_ascii=False)),
        "comparison": json.loads(comparison.to_json(orient="records", force_ascii=False)),
    }
    (output_dir / "metrics.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    backtest = results["backtest"]
    log = [
        "# 运行日志",
        "",
        f"- 命令: `python3 -m src.test.main`",
        f"- 样本区间: {backtest.index.min().date()} 至 {backtest.index.max().date()}",
        f"- 样本天数: {len(backtest)}",
        f"- 输出目录: `{output_dir}`",
        "- 数据: `~/local_data/financial_future_price.parquet`, `~/local_data/ashare_csiindex_trade.parquet`",
        "- 说明: 隔日反转因子采用本 PDF 可确认的 T-2 结算价涨跌幅近似。",
    ]
    (output_dir / "run_log.md").write_text("\n".join(log) + "\n", encoding="utf-8")


def main(config: BacktestConfig = CONFIG) -> None:
    """Execute the reproduction pipeline."""

    output_dir = config.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    results = build_backtest(config)
    _save_charts(results, output_dir)
    _save_excel_outputs(results, output_dir)
    _write_json_and_log(results, output_dir)

    summary = results["summary"]
    print("Backtest summary:")
    print(summary[["cumulative_return", "annual_return", "max_drawdown", "sharpe"]])
    print(f"Output saved to: {output_dir}")


if __name__ == "__main__":
    main()
