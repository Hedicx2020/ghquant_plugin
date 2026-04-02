"""
Common backtesting framework.

Provides:
- IC / RankIC calculation
- Quantile (group) backtesting
- Long-short portfolio backtesting
- Comprehensive performance analysis with visualisation and Excel output
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from scipy import stats

from common.utils import (
    calculate_annualized_return,
    calculate_annualized_volatility,
    calculate_max_drawdown,
    calculate_sharpe,
    calculate_win_rate,
    performance_summary,
)

# ---------------------------------------------------------------------------
# Matplotlib Chinese font setup
# ---------------------------------------------------------------------------
# Colour palette
BLUE = "#1f77b4"
RED = "#d62728"
CMAP = "RdBu_r"

import seaborn as sns
sns.set_style("whitegrid")

# Set Chinese font AFTER seaborn to avoid being overridden
plt.rcParams["font.sans-serif"] = ["Songti SC", "Heiti TC", "STHeiti", "Kaiti SC", "PingFang HK", "SimHei"]
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["font.family"] = "sans-serif"


# ---------------------------------------------------------------------------
# IC calculation
# ---------------------------------------------------------------------------

def calculate_ic(
    factor: pd.Series,
    forward_return: pd.Series,
) -> float:
    """Pearson IC between factor values and forward returns.

    Args:
        factor: Cross-sectional factor values.
        forward_return: Corresponding forward returns.

    Returns:
        Pearson correlation coefficient.
    """
    valid = pd.DataFrame({"f": factor, "r": forward_return}).dropna()
    if valid.shape[0] < 5:
        return np.nan
    return float(valid["f"].corr(valid["r"]))


def calculate_rank_ic(
    factor: pd.Series,
    forward_return: pd.Series,
) -> float:
    """Spearman RankIC between factor values and forward returns.

    Args:
        factor: Cross-sectional factor values.
        forward_return: Corresponding forward returns.

    Returns:
        Spearman rank correlation.
    """
    valid = pd.DataFrame({"f": factor, "r": forward_return}).dropna()
    if valid.shape[0] < 5:
        return np.nan
    corr, _ = stats.spearmanr(valid["f"], valid["r"])
    return float(corr)


def calculate_ic_series(
    factor_panel: pd.DataFrame,
    factor_col: str = "factor",
    return_col: str = "forward_return",
    date_col: str = "date",
    method: str = "rank",
) -> pd.DataFrame:
    """Calculate IC time-series.

    Args:
        factor_panel: Panel data with factor and forward return.
        factor_col: Factor column name.
        return_col: Forward return column name.
        date_col: Date column name.
        method: 'rank' for RankIC, 'pearson' for Pearson IC.

    Returns:
        DataFrame with columns [date, ic].
    """
    ic_func = calculate_rank_ic if method == "rank" else calculate_ic

    ic_records = (
        factor_panel
        .groupby(date_col)
        .apply(lambda g: ic_func(g[factor_col], g[return_col]), include_groups=False)
        .rename("ic")
        .reset_index()
    )
    return ic_records


def ic_summary(ic_series: pd.Series, periods_per_year: int = 12) -> dict:
    """Summary statistics for an IC series.

    Args:
        ic_series: Time-series of IC values.
        periods_per_year: Number of rebalancing periods per year (12=monthly, 52=weekly, 252=daily).

    Returns:
        Dict with ic_mean, ic_std, icir, win_rate, t_stat.
    """
    ic = ic_series.dropna()
    ic_mean = ic.mean()
    ic_std = ic.std()
    icir = ic_mean / ic_std * np.sqrt(periods_per_year) if ic_std > 0 else 0.0
    win_rate = (ic > 0).sum() / len(ic) if len(ic) > 0 else 0.0
    t_val = ic_mean / (ic_std / np.sqrt(len(ic))) if ic_std > 0 and len(ic) > 0 else 0.0
    return {
        "ic_mean": float(ic_mean),
        "ic_std": float(ic_std),
        "icir": float(icir),
        "win_rate": float(win_rate),
        "t_stat": float(t_val),
        "n_periods": len(ic),
    }


# ---------------------------------------------------------------------------
# Quantile (group) backtesting
# ---------------------------------------------------------------------------

def assign_quantile_groups(
    factor_panel: pd.DataFrame,
    factor_col: str = "factor",
    date_col: str = "date",
    n_groups: int = 5,
    group_col: str = "group",
) -> pd.DataFrame:
    """Assign stocks to quantile groups per cross-section.

    Group 1 = lowest factor, Group n = highest factor.

    Args:
        factor_panel: Panel data.
        factor_col: Factor column.
        date_col: Date column.
        n_groups: Number of groups.
        group_col: Output column name for group labels.

    Returns:
        DataFrame with added group column.
    """
    def _assign(g: pd.DataFrame) -> pd.Series:
        valid = g[factor_col].dropna()
        if valid.empty:
            return pd.Series(np.nan, index=g.index, name=group_col)
        labels = pd.qcut(valid.rank(method="first"), n_groups, labels=False) + 1
        return labels.reindex(g.index)

    factor_panel = factor_panel.copy()
    factor_panel[group_col] = (
        factor_panel.groupby(date_col, group_keys=False)
        .apply(_assign)
    )
    return factor_panel


def quantile_backtest(
    factor_panel: pd.DataFrame,
    factor_col: str = "factor",
    return_col: str = "forward_return",
    date_col: str = "date",
    n_groups: int = 5,
    transaction_cost: float = 0.003,
) -> dict:
    """Run quantile backtest.

    Args:
        factor_panel: Panel data with factor and forward return.
        factor_col: Factor column.
        return_col: Forward return column.
        date_col: Date column.
        n_groups: Number of groups.
        transaction_cost: Round-trip cost.

    Returns:
        Dictionary with group_returns, group_nav, turnover, etc.
    """
    panel = assign_quantile_groups(
        factor_panel, factor_col=factor_col, date_col=date_col, n_groups=n_groups
    )

    # Equal-weight group returns per period
    group_returns = (
        panel.dropna(subset=[return_col, "group"])
        .groupby([date_col, "group"])[return_col]
        .mean()
        .unstack("group")
        .sort_index()
    )
    group_returns.columns = [f"G{int(c)}" for c in group_returns.columns]

    # Long-short: top group - bottom group
    top_col, bot_col = f"G{n_groups}", "G1"
    group_returns["long_short"] = group_returns[top_col] - group_returns[bot_col]

    # Simple turnover estimation per group
    group_members = (
        panel.dropna(subset=["group"])
        .groupby([date_col, "group"])
        .apply(lambda g: set(g["stock_code"].values) if "stock_code" in g.columns else set(), include_groups=False)
        .unstack("group")
    )

    turnover_dict: dict[str, list[float]] = {}
    for g_col in group_returns.columns:
        if g_col == "long_short":
            continue
        g_idx = int(g_col[1:])
        if g_idx in group_members.columns:
            members_series = group_members[g_idx].dropna()
            turnovers = []
            for i in range(1, len(members_series)):
                prev, curr = members_series.iloc[i - 1], members_series.iloc[i]
                if prev and curr:
                    overlap = len(prev & curr)
                    total = max(len(prev), len(curr))
                    turnovers.append(1 - overlap / total if total > 0 else 0)
                else:
                    turnovers.append(0)
            turnover_dict[g_col] = turnovers

    # Deduct transaction cost (one-way cost = half round-trip * turnover)
    # For simplicity, deduct cost from group returns using average turnover
    group_returns_net = group_returns.copy()
    for g_col, tvs in turnover_dict.items():
        avg_turnover = np.mean(tvs) if tvs else 0.0
        group_returns_net[g_col] = group_returns[g_col] - transaction_cost * avg_turnover

    # Recalculate long-short after cost
    group_returns_net["long_short"] = group_returns_net[top_col] - group_returns_net[bot_col]

    # Cumulative NAV
    group_nav = (1 + group_returns_net).cumprod()

    return {
        "group_returns": group_returns,
        "group_returns_net": group_returns_net,
        "group_nav": group_nav,
        "turnover": turnover_dict,
        "panel": panel,
        "n_groups": n_groups,
    }


def long_short_backtest(
    factor_panel: pd.DataFrame,
    factor_col: str = "factor",
    return_col: str = "forward_return",
    date_col: str = "date",
    n_groups: int = 5,
    transaction_cost: float = 0.003,
    periods_per_year: int = 12,
) -> dict:
    """Long-short backtest: long top group, short bottom group.

    Args:
        factor_panel: Panel data.
        factor_col: Factor column.
        return_col: Forward return column.
        date_col: Date column.
        n_groups: Number of groups.
        transaction_cost: Round-trip cost.
        periods_per_year: Number of rebalancing periods per year (12=monthly, 52=weekly, 252=daily).

    Returns:
        Dictionary with returns, nav, and performance metrics.
    """
    qbt = quantile_backtest(
        factor_panel, factor_col, return_col, date_col, n_groups, transaction_cost
    )
    ls_returns = qbt["group_returns_net"]["long_short"]
    ls_nav = (1 + ls_returns).cumprod()

    perf = performance_summary(ls_returns, periods_per_year=periods_per_year, name="long_short")

    # Add top group metrics
    top_col = f"G{n_groups}"
    top_returns = qbt["group_returns_net"][top_col]
    top_perf = performance_summary(top_returns, periods_per_year=periods_per_year, name="top_group")

    return {
        "ls_returns": ls_returns,
        "ls_nav": ls_nav,
        "ls_performance": perf,
        "top_returns": top_returns,
        "top_performance": top_perf,
        "quantile_result": qbt,
    }


# ---------------------------------------------------------------------------
# Performance analysis & output
# ---------------------------------------------------------------------------

def performance_analysis(
    factor_panel: pd.DataFrame,
    factor_col: str = "factor",
    return_col: str = "forward_return",
    date_col: str = "date",
    n_groups: int = 5,
    transaction_cost: float = 0.003,
    output_dir: Optional[Path] = None,
    report_name: str = "factor",
    periods_per_year: int = 12,
) -> dict:
    """Full factor performance analysis pipeline.

    Computes IC series, quantile backtest, long-short returns,
    generates charts and Excel files.

    Args:
        factor_panel: Panel data with factor values and forward returns.
        factor_col: Factor column.
        return_col: Forward return column.
        date_col: Date column.
        n_groups: Number of groups.
        transaction_cost: Round-trip transaction cost.
        output_dir: Directory to save outputs.
        report_name: Report name for titles.
        periods_per_year: Number of rebalancing periods per year (12=monthly, 52=weekly, 252=daily).

    Returns:
        Dictionary with all computed results.
    """
    # --- IC analysis ---
    ic_df = calculate_ic_series(
        factor_panel, factor_col, return_col, date_col, method="rank"
    )
    ic_stats = ic_summary(ic_df["ic"], periods_per_year=periods_per_year)

    pearson_ic_df = calculate_ic_series(
        factor_panel, factor_col, return_col, date_col, method="pearson"
    )

    # --- Quantile backtest ---
    ls_result = long_short_backtest(
        factor_panel, factor_col, return_col, date_col, n_groups, transaction_cost,
        periods_per_year=periods_per_year,
    )
    qbt = ls_result["quantile_result"]
    group_returns_net = qbt["group_returns_net"]
    group_nav = qbt["group_nav"]

    # Per-group performance
    group_perf_records = []
    for col in group_returns_net.columns:
        perf = performance_summary(group_returns_net[col], periods_per_year=periods_per_year, name=col)
        # Add average turnover
        if col in qbt["turnover"]:
            perf["avg_turnover"] = float(np.mean(qbt["turnover"][col]))
        else:
            perf["avg_turnover"] = np.nan
        group_perf_records.append(perf)
    group_perf_df = pd.DataFrame(group_perf_records).set_index("name")

    results = {
        "ic_df": ic_df,
        "ic_stats": ic_stats,
        "pearson_ic_df": pearson_ic_df,
        "ls_result": ls_result,
        "group_returns_net": group_returns_net,
        "group_nav": group_nav,
        "group_perf_df": group_perf_df,
    }

    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        _save_charts(results, output_dir, report_name, n_groups)
        _save_excel(results, output_dir, report_name)

    return results


# ---------------------------------------------------------------------------
# Chart generation
# ---------------------------------------------------------------------------

def _save_charts(
    results: dict,
    output_dir: Path,
    report_name: str,
    n_groups: int,
) -> None:
    """Generate and save all required charts."""
    ic_df = results["ic_df"]
    group_nav = results["group_nav"]
    group_returns_net = results["group_returns_net"]
    group_perf_df = results["group_perf_df"]

    # 1. IC time-series (双轴: 柱状图 RankIC + 折线图 累计 RankIC)
    fig, ax1 = plt.subplots(figsize=(14, 6))

    ic_values = ic_df["ic"].values
    dates = ic_df["date"].values
    n = len(dates)

    # 左轴: RankIC 柱状图（用数字索引避免日期间距不均）
    x_idx = np.arange(n)
    bar_colors = [BLUE if v >= 0 else RED for v in ic_values]
    ax1.bar(x_idx, ic_values, color=bar_colors, alpha=0.7, width=0.8, label="RankIC")
    ax1.axhline(ic_df["ic"].mean(), color="gray", linestyle="--", linewidth=1.2,
                label=f'IC均值={ic_df["ic"].mean():.4f}')
    ax1.axhline(0, color="black", linewidth=0.5)
    ax1.set_ylabel("RankIC", fontsize=12)
    ax1.tick_params(axis="y")

    # x 轴标签：每隔 ~24 期标一个年份
    tick_step = max(1, n // 8)
    ax1.set_xticks(x_idx[::tick_step])
    ax1.set_xticklabels([pd.Timestamp(d).strftime("%Y-%m") for d in dates[::tick_step]],
                        rotation=45, fontsize=9)

    # 右轴: 累计 RankIC 折线图
    ax2 = ax1.twinx()
    cum_ic = ic_df["ic"].cumsum()
    ax2.plot(x_idx, cum_ic.values, color="darkorange", linewidth=2.5, label="累计RankIC")
    ax2.set_ylabel("累计RankIC", fontsize=12, color="darkorange")
    ax2.tick_params(axis="y", labelcolor="darkorange")

    # 合并图例
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, fontsize=10, loc="upper left")

    ax1.set_title(f"{report_name} RankIC 时间序列", fontsize=14)
    ax1.grid(True, alpha=0.3, axis="y")
    ax1.set_xlim(-1, n)
    fig.tight_layout()
    fig.savefig(output_dir / "ic_series.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 2. IC distribution histogram
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.hist(ic_df["ic"].dropna(), bins=30, color=BLUE, alpha=0.7, edgecolor="white")
    ax.axvline(ic_df["ic"].mean(), color=RED, linestyle="--", linewidth=1.5,
               label=f'均值={ic_df["ic"].mean():.4f}')
    ax.set_title(f"{report_name} RankIC 分布", fontsize=14)
    ax.set_xlabel("RankIC", fontsize=12)
    ax.set_ylabel("频数", fontsize=12)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(output_dir / "ic_distribution.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 3. Group cumulative returns
    fig, ax = plt.subplots(figsize=(12, 6))
    cmap = plt.get_cmap(CMAP)
    group_cols = [c for c in group_nav.columns if c.startswith("G")]
    colors = [cmap(i / (len(group_cols) - 1)) for i in range(len(group_cols))]
    for col, color in zip(group_cols, colors):
        ax.plot(group_nav.index, group_nav[col], label=col, color=color, linewidth=1.5)
    if "long_short" in group_nav.columns:
        ax.plot(group_nav.index, group_nav["long_short"], label="多空",
                color="black", linewidth=2, linestyle="--")
    ax.set_title(f"{report_name} 分组累计收益", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("累计净值", fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(output_dir / "group_cumulative_returns.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 4. Group returns bar chart
    fig, ax = plt.subplots(figsize=(10, 6))
    g_cols = [c for c in group_perf_df.index if c.startswith("G")]
    ann_rets = group_perf_df.loc[g_cols, "ann_return"]
    bar_colors = [BLUE if i >= len(g_cols) // 2 else RED for i in range(len(g_cols))]
    ax.bar(ann_rets.index, ann_rets.values * 100, color=bar_colors, alpha=0.8)
    ax.set_title(f"{report_name} 分组年化收益", fontsize=14)
    ax.set_xlabel("分组", fontsize=12)
    ax.set_ylabel("年化收益 (%)", fontsize=12)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f%%"))
    fig.tight_layout()
    fig.savefig(output_dir / "group_returns_bar.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 5. Net value comparison (top group vs long-short vs benchmark)
    fig, ax = plt.subplots(figsize=(12, 6))
    top_col = f"G{n_groups}"
    if top_col in group_nav.columns:
        ax.plot(group_nav.index, group_nav[top_col], label=f"多头({top_col})",
                color=BLUE, linewidth=1.5)
    if "long_short" in group_nav.columns:
        ax.plot(group_nav.index, group_nav["long_short"], label="多空组合",
                color=RED, linewidth=1.5)
    bot_col = "G1"
    if bot_col in group_nav.columns:
        ax.plot(group_nav.index, group_nav[bot_col], label=f"空头({bot_col})",
                color="gray", linewidth=1.0, linestyle="--")
    ax.set_title(f"{report_name} 净值对比", fontsize=14)
    ax.set_xlabel("日期", fontsize=12)
    ax.set_ylabel("净值", fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(output_dir / "net_value_comparison.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


def _save_excel(
    results: dict,
    output_dir: Path,
    report_name: str,
) -> None:
    """Save all results to a single consolidated Excel file with charts embedded."""
    import io
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    excel_path = output_dir / f"{report_name}_results.xlsx"

    ic_df = results["ic_df"]
    ic_stats = results["ic_stats"]
    group_returns_net = results["group_returns_net"]
    group_nav = results["group_nav"]
    group_perf_df = results["group_perf_df"]

    # --- Write data using pandas ExcelWriter, then embed images ---
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Sheet 1: Backtest Summary
        summary_data = {
            "指标": [
                "RankIC均值", "RankIC标准差", "RankICIR(年化)",
                "RankIC胜率", "T统计量", "样本期数",
            ],
            "值": [
                f"{ic_stats['ic_mean']:.4f}",
                f"{ic_stats['ic_std']:.4f}",
                f"{ic_stats['icir']:.2f}",
                f"{ic_stats['win_rate']:.2%}",
                f"{ic_stats['t_stat']:.2f}",
                str(ic_stats["n_periods"]),
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="回测摘要", index=False, startrow=0)

        # Group performance
        gp = group_perf_df.copy()
        gp.index.name = "分组"
        gp.to_excel(writer, sheet_name="回测摘要", startrow=len(summary_data["指标"]) + 3)

        # Sheet 2: IC Series + chart
        ic_out = ic_df.copy()
        ic_out.columns = ["日期", "RankIC"]
        ic_out.to_excel(writer, sheet_name="IC序列", index=False)

        # Sheet 3: Group Returns + chart
        gr = group_returns_net.copy()
        gr.index.name = "日期"
        gr.to_excel(writer, sheet_name="分组收益")

        # Sheet 4: Group NAV + chart
        gn = group_nav.copy()
        gn.index.name = "日期"
        gn.to_excel(writer, sheet_name="分组净值")

    # --- Re-open and embed images ---
    wb = load_workbook(excel_path)

    def _embed_image(ws, img_path: Path, anchor: str = "H2") -> None:
        if img_path.exists():
            img = XlImage(str(img_path))
            img.width = 720
            img.height = 360
            ws.add_image(img, anchor)

    # Bold header row
    bold_font = Font(bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = bold_font
        ws.sheet_properties.tabColor = "1F77B4"
        # Auto-width
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)
        ws.freeze_panes = "A2"

    # Embed charts next to data
    _embed_image(wb["IC序列"], output_dir / "ic_series.png", "D2")
    _embed_image(wb["IC序列"], output_dir / "ic_distribution.png", "D22")
    _embed_image(wb["分组收益"], output_dir / "group_returns_bar.png", "J2")
    _embed_image(wb["分组净值"], output_dir / "group_cumulative_returns.png", "J2")
    _embed_image(wb["分组净值"], output_dir / "net_value_comparison.png", "J22")

    wb.save(excel_path)
    print(f"Excel saved: {excel_path}")
