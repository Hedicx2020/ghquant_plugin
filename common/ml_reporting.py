"""机器学习研报的通用结果汇总、证据明细、Excel 与图表输出层。

本模块只消费已经完成训练和回测的结构化对象，不参与模型拟合、组合构造或
指标裁判。研报摘要基准与复现计算值始终分栏保存，避免把原文数字写成结果。
"""

from __future__ import annotations

import json
import math
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from common.ml_portfolio import MLPortfolioResult


def _jsonable(value: Any) -> Any:
    """递归转换为严格 JSON 值，并把 NaN/Inf 置为 null。"""
    if isinstance(value, Mapping):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_jsonable(item) for item in value]
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        timestamp = pd.Timestamp(value)
        return None if pd.isna(timestamp) else timestamp.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if value is pd.NA or (not isinstance(value, (str, bytes)) and pd.isna(value)):
        return None
    return value


def write_json(path: Path, payload: Any, *, indent: int) -> None:
    """以 UTF-8 和严格 JSON 原子落盘。"""
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    temporary.write_text(
        json.dumps(_jsonable(payload), ensure_ascii=False, indent=indent, allow_nan=False),
        encoding="utf-8",
    )
    temporary.replace(destination)


def calculate_gradient_feature_importance(
    models: Mapping[str, Any],
    dataset: Sequence[Any],
    feature_names: Sequence[str],
    *,
    batch_size: int,
) -> pd.DataFrame:
    """分批遍历完整数据集，累计实际模型输出对输入的绝对梯度。"""
    import torch

    if batch_size <= 0 or len(dataset) <= 0:
        raise ValueError("特征重要性 batch_size 与数据集长度必须为正")
    records: list[dict[str, Any]] = []
    for model_name, model in models.items():
        model = model.to("cpu")
        model.eval()
        gradient_sum = np.zeros(len(feature_names), dtype=np.float64)
        observation_count = 0
        time_steps: int | None = None
        for start in range(0, len(dataset), batch_size):
            stop = min(start + batch_size, len(dataset))
            features = np.stack([dataset[index][0] for index in range(start, stop)]).astype(
                np.float32, copy=False
            )
            if features.ndim != 3 or features.shape[-1] != len(feature_names):
                raise ValueError("特征重要性批次必须为 [batch,time,features] 且通道数一致")
            if time_steps is None:
                time_steps = int(features.shape[1])
            elif time_steps != int(features.shape[1]):
                raise ValueError("完整回测数据集的特征窗口长度不一致")
            model.zero_grad(set_to_none=True)
            inputs = torch.tensor(features, dtype=torch.float32, requires_grad=True)
            model(inputs).reshape(-1).sum().backward()
            if inputs.grad is None or not torch.isfinite(inputs.grad).all():
                raise FloatingPointError(f"{model_name} 输入梯度不可用或包含非有限值")
            gradient_sum += (
                inputs.grad.detach().abs().sum(dim=(0, 1)).cpu().numpy().astype(np.float64)
            )
            observation_count += int(features.shape[0] * features.shape[1])
        raw = gradient_sum / observation_count
        total = float(raw.sum())
        normalized = raw / total if total > 0.0 else np.full_like(raw, np.nan)
        records.extend(
            {
                "model": str(model_name),
                "feature": str(feature),
                "mean_absolute_input_gradient": float(raw_value),
                "normalized_importance": float(normalized_value),
                "method": "mean_absolute_input_gradient",
                "sample_count": int(len(dataset)),
                "time_steps": int(time_steps or 0),
                "batch_size": int(batch_size),
                "full_dataset": True,
            }
            for feature, raw_value, normalized_value in zip(feature_names, raw, normalized)
        )
    return pd.DataFrame(records)


def calculate_model_factor_correlation(
    factor_panel: pd.DataFrame,
    *,
    model_names: Sequence[str],
    method: str,
    min_common: int,
    min_unique_values: int,
    sample_scope: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """按 AS18 对10个无序模型对计算成对完整样本 Pearson。"""
    if method != "pearson":
        raise ValueError("AS18 当前只允许 Pearson 模型间相关性")
    required = {"model", "stock_code", "date", "prediction"}
    missing = sorted(required - set(factor_panel.columns))
    if missing:
        raise KeyError(f"模型相关性面板缺少字段: {missing}")
    wide = factor_panel.pivot(
        index=["date", "stock_code"], columns="model", values="prediction"
    ).reindex(columns=list(model_names))
    pair_records: list[dict[str, Any]] = []
    matrix = pd.DataFrame(np.eye(len(model_names)), index=model_names, columns=model_names)
    for left_index, model_left in enumerate(model_names):
        for model_right in model_names[left_index + 1 :]:
            pair = wide.loc[:, [model_left, model_right]].replace([np.inf, -np.inf], np.nan).dropna()
            n_common = int(len(pair))
            left_unique = int(pair[model_left].nunique())
            right_unique = int(pair[model_right].nunique())
            if n_common < min_common:
                correlation = np.nan
                status = "insufficient_common_samples"
            elif min(left_unique, right_unique) < min_unique_values:
                correlation = np.nan
                status = "constant_prediction"
            else:
                correlation = float(pair[model_left].corr(pair[model_right], method="pearson"))
                status = "valid" if np.isfinite(correlation) else "non_finite"
            matrix.loc[model_left, model_right] = correlation
            matrix.loc[model_right, model_left] = correlation
            pair_records.append(
                {
                    "model_left": model_left,
                    "model_right": model_right,
                    "correlation": correlation,
                    "n_common": n_common,
                    "left_unique_values": left_unique,
                    "right_unique_values": right_unique,
                    "method": method,
                    "sample_scope": sample_scope,
                    "min_common": min_common,
                    "status": status,
                }
            )
    matrix.index.name = "model_left"
    matrix_long = matrix.reset_index().melt(
        id_vars="model_left", var_name="model_right", value_name="correlation"
    )
    daily_records: list[dict[str, Any]] = []
    for signal_date, daily_wide in wide.groupby(level="date", sort=True):
        daily_wide = daily_wide.droplevel("date")
        for left_index, model_left in enumerate(model_names):
            for model_right in model_names[left_index + 1 :]:
                pair = daily_wide.loc[:, [model_left, model_right]].replace(
                    [np.inf, -np.inf], np.nan
                ).dropna()
                n_common = int(len(pair))
                left_unique = int(pair[model_left].nunique())
                right_unique = int(pair[model_right].nunique())
                if n_common < min_common:
                    correlation = np.nan
                    status = "insufficient_common_samples"
                elif min(left_unique, right_unique) < min_unique_values:
                    correlation = np.nan
                    status = "constant_prediction"
                else:
                    correlation = float(pair[model_left].corr(pair[model_right], method="pearson"))
                    status = "valid" if np.isfinite(correlation) else "non_finite"
                daily_records.append(
                    {
                        "date": signal_date,
                        "model_left": model_left,
                        "model_right": model_right,
                        "correlation": correlation,
                        "n_common": n_common,
                        "left_unique_values": left_unique,
                        "right_unique_values": right_unique,
                        "method": method,
                        "sample_scope": f"daily_{sample_scope}",
                        "min_common": min_common,
                        "status": status,
                    }
                )
    return matrix_long, pd.DataFrame(pair_records), pd.DataFrame(daily_records)


def _off_diagonal_correlation_range(pair_summary: pd.DataFrame) -> tuple[float | None, float | None]:
    finite = pair_summary.loc[
        pair_summary["status"].eq("valid"), "correlation"
    ].to_numpy(
        dtype=float
    )
    finite = finite[np.isfinite(finite)]
    if finite.size == 0:
        return None, None
    return float(finite.min()), float(finite.max())


def _rank_ic_summary(
    path_aggregate: pd.DataFrame,
    path_summary: pd.DataFrame,
) -> pd.DataFrame:
    """以20条路径等权聚合为主值，并附逐路径均值的离散度。"""
    path_distribution = (
        path_summary.groupby("model", observed=True, as_index=False)
        .agg(
            rank_ic_path_mean_std=("rank_ic_time_mean", "std"),
            rank_ic_path_mean_median=("rank_ic_time_mean", "median"),
            positive_path_rate=(
                "rank_ic_time_mean",
                lambda values: float((values > 0).mean()),
            ),
        )
    )
    return path_aggregate.merge(
        path_distribution,
        on="model",
        how="left",
        validate="one_to_one",
    )


def _label_coverage_tables(
    portfolio: MLPortfolioResult,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """复用 m4 标签诊断字段，形成模型摘要与逐日覆盖证据。"""
    count_columns = [
        "candidate_count",
        "valid_label_count",
        "label_missing_count",
        "valid_label_rate",
        "label_missing_rate",
    ]
    daily_columns = ["model", "date", "rank_ic", *count_columns]
    summary_columns = ["model", "signal_count", *count_columns]
    daily_missing = sorted(set(daily_columns) - set(portfolio.rank_ic.columns))
    summary_missing = sorted(
        set(summary_columns) - set(portfolio.rank_ic_path_aggregate.columns)
    )
    if daily_missing or summary_missing:
        raise KeyError(
            "m4 标签覆盖字段不完整: "
            f"daily_missing={daily_missing}, summary_missing={summary_missing}"
        )
    daily = portfolio.rank_ic.loc[:, daily_columns].copy()
    summary = portfolio.rank_ic_path_aggregate.loc[:, summary_columns].copy()
    for frame in (daily, summary):
        if not (
            frame["candidate_count"]
            == frame["valid_label_count"] + frame["label_missing_count"]
        ).all():
            raise AssertionError("标签覆盖计数不满足 candidate=valid+missing")
    summary["portfolio_candidates_include_missing_labels"] = True
    summary["rank_ic_pairwise_complete_labels"] = True
    summary["prediction_scatter_pairwise_complete_labels"] = True
    return summary, daily


def build_report_reference_tables(
    *,
    report_gpu_memory_gb: Sequence[tuple[str, float]],
    report_performance: Sequence[tuple[str, str, str, float, float]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """把规格书摘要构造成只读对照表，不进入复现指标计算。"""
    resource_reference = pd.DataFrame(
        report_gpu_memory_gb, columns=["model", "report_gpu_memory_gb"]
    )
    performance_reference = pd.DataFrame(
        report_performance,
        columns=["model", "portfolio", "period", "report_annualized_return", "report_sharpe"],
    )
    return resource_reference, performance_reference


def build_ml_metrics(
    *,
    run_metadata: Mapping[str, Any],
    sample_coverage: pd.DataFrame,
    resource_summary: pd.DataFrame,
    loss_curves: pd.DataFrame,
    portfolio: MLPortfolioResult,
    factor_correlation: pd.DataFrame,
    factor_correlation_pairs: pd.DataFrame,
    factor_correlation_by_date: pd.DataFrame,
    feature_importance: pd.DataFrame,
    report_ltc_improvement_range: tuple[float, float],
    report_factor_correlation_range: tuple[float, float],
    report_resource_reference: pd.DataFrame,
    report_performance_reference: pd.DataFrame,
    output_boundaries: Mapping[str, Any],
) -> tuple[dict[str, Any], pd.DataFrame]:
    """从完整训练/回测明细构建 metrics 与非裁判型 comparison 表。"""
    correlation_min, correlation_max = _off_diagonal_correlation_range(
        factor_correlation_pairs
    )
    rank_ic_summary = _rank_ic_summary(
        portfolio.rank_ic_path_aggregate,
        portfolio.rank_ic_path_summary,
    )
    label_coverage_summary, label_coverage_by_date = _label_coverage_tables(portfolio)
    reproduced = portfolio.path_metric_aggregate.loc[
        portfolio.path_metric_aggregate["portfolio"].isin(["top_long", "top_bottom"])
    ].copy()
    comparison = report_performance_reference.merge(
        reproduced,
        on=["model", "portfolio", "period"],
        how="left",
        validate="one_to_one",
    )
    comparison["annualized_return_deviation"] = (
        comparison["annualized_return"] - comparison["report_annualized_return"]
    )
    comparison["sharpe_deviation"] = comparison["sharpe"] - comparison["report_sharpe"]
    comparison["element_id"] = comparison.apply(
        lambda row: {
            ("top_long", "all"): "R4",
            ("top_long", "2024"): "R5",
            ("top_bottom", "all"): "R6",
            ("top_bottom", "2024"): "R7",
        }[(str(row["portfolio"]), str(row["period"]))],
        axis=1,
    )
    comparison["calculation_scope"] = "完整复现样本；逐路径指标后算术平均"
    comparison["verdict"] = "pending_verifier"

    resource = report_resource_reference.merge(
        resource_summary, on="model", how="left", validate="one_to_one"
    )
    resource["comparison_status"] = "not_comparable_hardware_batch_dtype_unknown_in_report"
    resource["report_value_used_as_reproduction"] = False

    all_period = portfolio.path_metric_aggregate.loc[
        portfolio.path_metric_aggregate["period"].eq("all")
    ].copy()
    annual = portfolio.path_metric_aggregate.loc[
        ~portfolio.path_metric_aggregate["period"].eq("all")
    ].copy()
    elements = {
        "R1": {
            "status": "reference_only",
            "report_literature_improvement_range": report_ltc_improvement_range,
            "reproduced_value": None,
            "reason": "文献背景摘要，不把5%至70%抄作本次模型性能",
        },
        "R2": {
            "status": "computed_not_hardware_comparable",
            "report_reference": resource.to_dict("records"),
            "reproduced_resource_measurement": resource_summary.to_dict("records"),
        },
        "R3": {
            "status": "computed",
            "report_factor_correlation_range": report_factor_correlation_range,
            "reproduced_off_diagonal_range": [correlation_min, correlation_max],
            "matrix": factor_correlation.to_dict("records"),
            "unordered_pairs": factor_correlation_pairs.to_dict("records"),
        },
        "R4": comparison.loc[comparison["element_id"].eq("R4")].to_dict("records"),
        "R5": comparison.loc[comparison["element_id"].eq("R5")].to_dict("records"),
        "R6": comparison.loc[comparison["element_id"].eq("R6")].to_dict("records"),
        "R7": comparison.loc[comparison["element_id"].eq("R7")].to_dict("records"),
        "SA1": {
            "status": "computed",
            "scope": "2024子样本与全区间并列",
            "rows": comparison.loc[comparison["period"].isin(["all", "2024"])].to_dict("records"),
        },
        "SA2": {
            "status": "computed_non_original_table",
            "off_diagonal_range": [correlation_min, correlation_max],
            "overall_unordered_pairs": factor_correlation_pairs.to_dict("records"),
            "daily_pair_diagnostics": factor_correlation_by_date.to_dict("records"),
            "original_table_status": "reference_only_data_missing",
        },
        "SA3": {
            "status": "computed_non_original_table",
            "annual_performance": annual.to_dict("records"),
            "original_tables_status": "TBL6-TBL10_reference_only_data_missing",
        },
    }
    metrics = {
        "schema_version": 1,
        "report_id": "r002_liquid_neural_networks",
        "run_metadata": dict(run_metadata),
        "method_scope": "方法级降级复现；固定开源标准实现，非原研报私有参数",
        "data_missing_boundary": {
            "original_figures": "FIG15-FIG24 reference_only/data_missing",
            "original_tables": "TBL4-TBL10逐格内容reference_only/data_missing",
            "own_charts": "全部标题含复现生成—非原图复刻，不沿用原编号或题名",
        },
        "output_boundaries": dict(output_boundaries),
        "sample_coverage": sample_coverage.to_dict("records"),
        "training": {
            "loss_curves": loss_curves.to_dict("records"),
            "resources": resource_summary.to_dict("records"),
        },
        "prediction_evaluation": {
            "rank_ic": rank_ic_summary.to_dict("records"),
            "label_coverage_summary": label_coverage_summary.to_dict("records"),
            "label_coverage_by_date": label_coverage_by_date.to_dict("records"),
            "label_coverage_policy": {
                "portfolio_candidates": "all_valid_predictions_including_missing_labels",
                "rank_ic": "pairwise_complete_prediction_and_label",
                "prediction_scatter": "pairwise_complete_prediction_and_label_before_sampling",
                "label_missing_changes_groups_or_trades": False,
            },
            "factor_correlation_off_diagonal_range": [correlation_min, correlation_max],
            "factor_correlation_pairs": factor_correlation_pairs.to_dict("records"),
            "factor_correlation_by_date": factor_correlation_by_date.to_dict("records"),
            "monotonicity": portfolio.monotonicity.to_dict("records"),
            "feature_importance": feature_importance.to_dict("records"),
        },
        "portfolio_evaluation": {
            "full_period": all_period.to_dict("records"),
            "annual": annual.to_dict("records"),
        },
        "elements": elements,
        "verifier_contract": {
            "comparison_source": "elements.R1-R7 与完整Parquet明细",
            "verdict": "pending_verifier",
            "report_values_are_not_reproduced_values": True,
        },
    }
    comparison_columns = [
        "element_id",
        "model",
        "portfolio",
        "period",
        "report_annualized_return",
        "annualized_return",
        "annualized_return_deviation",
        "report_sharpe",
        "sharpe",
        "sharpe_deviation",
        "path_count",
        "n_days_mean",
        "calculation_scope",
        "verdict",
    ]
    return metrics, comparison.loc[:, comparison_columns]


def _setup_plot_theme() -> None:
    sns.set_theme(style="whitegrid")
    plt.rcParams["font.sans-serif"] = [
        "PingFang SC",
        "Arial Unicode MS",
        "Heiti TC",
        "SimHei",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def _save_figure(figure: plt.Figure, path: Path, dpi: int) -> None:
    figure.tight_layout()
    figure.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(figure)


def _deterministic_display_sample(frame: pd.DataFrame, limit: int, seed: int) -> pd.DataFrame:
    if len(frame) <= limit:
        return frame.copy()
    return frame.sample(n=limit, random_state=seed).sort_index().copy()


def build_ml_chart_data(
    *,
    portfolio: MLPortfolioResult,
    factor_correlation: pd.DataFrame,
    loss_curves: pd.DataFrame,
    resource_summary: pd.DataFrame,
    feature_importance: pd.DataFrame,
    plot_point_limit: int,
    seed: int,
) -> dict[str, pd.DataFrame]:
    """一次构造8张图的唯一数据对象，供绘图与Excel原样复用。"""
    monthly_ic = portfolio.rank_ic_path_series.copy()
    monthly_ic["month"] = pd.to_datetime(monthly_ic["date"]).dt.to_period("M").dt.to_timestamp()
    monthly_path_ic = (
        monthly_ic.groupby(
            ["model", "path_offset", "month"], observed=True, as_index=False
        )
        .agg(
            rank_ic=("rank_ic", "mean"),
            candidate_count=("candidate_count", "sum"),
            valid_label_count=("valid_label_count", "sum"),
            label_missing_count=("label_missing_count", "sum"),
        )
    )
    monthly_ic = (
        monthly_path_ic
        .groupby(["model", "month"], observed=True, as_index=False)
        .agg(
            rank_ic=("rank_ic", "mean"),
            active_path_count=("path_offset", "nunique"),
            candidate_count=("candidate_count", "sum"),
            valid_label_count=("valid_label_count", "sum"),
            label_missing_count=("label_missing_count", "sum"),
        )
    )
    monthly_ic["valid_label_rate"] = (
        monthly_ic["valid_label_count"] / monthly_ic["candidate_count"]
    )
    monthly_ic["label_missing_rate"] = (
        monthly_ic["label_missing_count"] / monthly_ic["candidate_count"]
    )
    group_nav = (
        portfolio.path_nav.groupby(["model", "group", "date"], observed=True, as_index=False)
        .agg(nav=("nav", "mean"), active_path_count=("path_offset", "nunique"))
        .sort_values(["model", "group", "date"])
    )
    group_nav["cumulative_return"] = group_nav["nav"] - 1.0
    group_display = group_nav.loc[group_nav["group"].isin([1, 5, 10])].reset_index(drop=True)
    scatter_source = portfolio.factor_panel.loc[
        :, ["model", "stock_code", "date", "prediction", "label"]
    ].dropna(subset=["prediction", "label"])
    scatter = _deterministic_display_sample(
        scatter_source,
        plot_point_limit,
        seed,
    ).reset_index(drop=True)
    annual = portfolio.path_metric_aggregate.loc[
        portfolio.path_metric_aggregate["portfolio"].isin(["top_long", "top_bottom"])
        & ~portfolio.path_metric_aggregate["period"].eq("all")
    ].copy()
    resources = resource_summary.copy()
    has_device_measurement = resources["device_memory_incremental_peak_bytes"].notna()
    resources["display_memory_bytes"] = resources["device_memory_incremental_peak_bytes"].where(
        has_device_measurement, resources["cpu_process_peak_rss_bytes"]
    )
    resources["display_memory_scope"] = np.where(
        has_device_measurement,
        resources["device_memory_scope"],
        resources["cpu_memory_measurement"],
    )
    # GRU 上游没有 wiring 字段；四个液态模型则必须显式报告布尔值，
    # 防止把未知液态 wiring 静默归作 dense。
    expected_models = {"GRU", "LTC", "NCP（LTC）", "CFC", "NCP（CFC）"}
    actual_models = set(resources["model"].astype(str))
    if actual_models != expected_models or resources["model"].duplicated().any():
        raise ValueError("资源表必须且只能包含五个唯一的约定模型")
    sparse_wiring = resources["is_sparse_wiring"]
    missing_wiring_models = set(resources.loc[sparse_wiring.isna(), "model"].astype(str))
    if missing_wiring_models - {"GRU"}:
        raise ValueError(
            f"液态模型缺少 is_sparse_wiring: {sorted(missing_wiring_models - {'GRU'})}"
        )
    invalid_wiring_rows = [
        index
        for index, value in sparse_wiring.items()
        if pd.notna(value) and not isinstance(value, (bool, np.bool_))
    ]
    if invalid_wiring_rows:
        raise ValueError("is_sparse_wiring 只允许布尔值或 GRU 缺失值")
    resources["is_sparse_wiring"] = sparse_wiring.fillna(False).map(
        {True: "sparse", False: "dense"}
    )
    return {
        "feature_importance": feature_importance.reset_index(drop=True).copy(),
        "ic_series": monthly_ic.reset_index(drop=True),
        "group_cumulative_returns": group_display,
        "train_val_loss": loss_curves.reset_index(drop=True).copy(),
        "prediction_scatter": scatter,
        "model_correlation": factor_correlation.reset_index(drop=True).copy(),
        "annual_returns": annual.reset_index(drop=True),
        "resource_usage": resources.reset_index(drop=True),
    }


def create_ml_charts(
    *,
    output_dir: Path,
    chart_data: Mapping[str, pd.DataFrame],
    title_prefix: str,
    dpi: int,
    time_series_size: tuple[float, float],
    standard_size: tuple[float, float],
    primary_blue: str,
    primary_red: str,
    neutral_gray: str,
) -> dict[str, Path]:
    """生成 ml 必需五图及相关性、年度、资源补充图。"""
    _setup_plot_theme()
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    figure, axis = plt.subplots(figsize=standard_size)
    sns.barplot(
        data=chart_data["feature_importance"],
        x="feature",
        y="normalized_importance",
        hue="model",
        ax=axis,
    )
    axis.set_title(f"{title_prefix}：六通道输入梯度重要性", fontsize=14)
    axis.set_xlabel("特征", fontsize=12)
    axis.set_ylabel("归一化绝对输入梯度", fontsize=12)
    axis.tick_params(axis="x", rotation=25)
    paths["feature_importance.png"] = destination / "feature_importance.png"
    _save_figure(figure, paths["feature_importance.png"], dpi)

    figure, axis = plt.subplots(figsize=time_series_size)
    sns.lineplot(data=chart_data["ic_series"], x="month", y="rank_ic", hue="model", ax=axis)
    axis.axhline(0.0, color="black", linewidth=0.8)
    axis.set_title(f"{title_prefix}：预测因子月度RankIC", fontsize=14)
    axis.set_xlabel("月份", fontsize=12)
    axis.set_ylabel("RankIC", fontsize=12)
    paths["ic_series.png"] = destination / "ic_series.png"
    _save_figure(figure, paths["ic_series.png"], dpi)

    group_display = chart_data["group_cumulative_returns"]
    figure, axes = plt.subplots(len(group_display["model"].unique()), 1, figsize=(12, 18), sharex=True)
    axes = np.atleast_1d(axes)
    group_palette = {1: primary_blue, 5: neutral_gray, 10: primary_red}
    for axis, (model, frame) in zip(axes, group_display.groupby("model", sort=True)):
        sns.lineplot(
            data=frame,
            x="date",
            y="cumulative_return",
            hue="group",
            palette=group_palette,
            ax=axis,
        )
        axis.set_title(f"{title_prefix}：{model}")
        axis.set_xlabel("日期")
        axis.set_ylabel("累计收益率")
    figure.suptitle(f"{title_prefix}：十分档代表组累计收益", fontsize=14, y=1.002)
    paths["group_cumulative_returns.png"] = destination / "group_cumulative_returns.png"
    _save_figure(figure, paths["group_cumulative_returns.png"], dpi)

    figure, axis = plt.subplots(figsize=time_series_size)
    loss_curves = chart_data["train_val_loss"]
    sns.lineplot(data=loss_curves, x="epoch", y="train_loss", hue="model", ax=axis)
    for model, frame in loss_curves.groupby("model", sort=True):
        axis.plot(frame["epoch"], frame["validation_loss"], linestyle="--", label=f"{model} 验证")
    axis.set_title(f"{title_prefix}：五模型训练与验证损失", fontsize=14)
    axis.set_xlabel("Epoch", fontsize=12)
    axis.set_ylabel("MSE", fontsize=12)
    axis.legend(fontsize=8, ncol=2)
    paths["train_val_loss.png"] = destination / "train_val_loss.png"
    _save_figure(figure, paths["train_val_loss.png"], dpi)

    scatter = chart_data["prediction_scatter"]
    figure, axis = plt.subplots(figsize=standard_size)
    sns.scatterplot(
        data=scatter,
        x="prediction",
        y="label",
        hue="model",
        s=10,
        alpha=0.28,
        ax=axis,
    )
    axis.set_title(f"{title_prefix}：样本外预测与实际20日收益", fontsize=14)
    axis.set_xlabel("预测收益", fontsize=12)
    axis.set_ylabel("实际收益", fontsize=12)
    paths["prediction_scatter.png"] = destination / "prediction_scatter.png"
    _save_figure(figure, paths["prediction_scatter.png"], dpi)

    correlation_matrix = chart_data["model_correlation"].pivot(
        index="model_left", columns="model_right", values="correlation"
    )
    figure, axis = plt.subplots(figsize=standard_size)
    sns.heatmap(correlation_matrix, annot=True, fmt=".3f", cmap="RdBu_r", center=0.0, ax=axis)
    axis.set_title(f"{title_prefix}：五模型样本外因子相关性", fontsize=14)
    axis.set_xlabel("模型", fontsize=12)
    axis.set_ylabel("模型", fontsize=12)
    paths["model_correlation.png"] = destination / "model_correlation.png"
    _save_figure(figure, paths["model_correlation.png"], dpi)

    annual = chart_data["annual_returns"]
    figure, axes = plt.subplots(2, 1, figsize=(12, 10), sharex=True)
    for axis, portfolio_name in zip(axes, ["top_long", "top_bottom"]):
        frame = annual.loc[annual["portfolio"].eq(portfolio_name)]
        sns.lineplot(data=frame, x="period", y="annualized_return", hue="model", marker="o", ax=axis)
        axis.axhline(0.0, color="black", linewidth=0.8)
        portfolio_label = "多头" if portfolio_name == "top_long" else "多空"
        axis.set_title(f"{title_prefix}：{portfolio_label}")
        axis.set_ylabel("年化收益率")
    axes[-1].set_xlabel("年份")
    figure.suptitle(f"{title_prefix}：五模型分年度组合表现", fontsize=14)
    paths["annual_returns.png"] = destination / "annual_returns.png"
    _save_figure(figure, paths["annual_returns.png"], dpi)

    figure, axes = plt.subplots(1, 2, figsize=(13, 6))
    resource_summary = chart_data["resource_usage"]
    sns.barplot(data=resource_summary, x="model", y="parameter_count", color=primary_blue, ax=axes[0])
    axes[0].set_title(f"{title_prefix}：模型参数量")
    axes[0].tick_params(axis="x", rotation=25)
    axes[0].set_xlabel("模型")
    axes[0].set_ylabel("参数个数")
    memory_column = "display_memory_bytes"
    sns.barplot(data=resource_summary, x="model", y=memory_column, color=primary_red, ax=axes[1])
    memory_label = "隔离设备增量峰值；无设备值时为累计CPU RSS"
    axes[1].set_title(f"{title_prefix}：{memory_label}")
    axes[1].tick_params(axis="x", rotation=25)
    axes[1].set_xlabel("模型")
    axes[1].set_ylabel("字节")
    figure.suptitle(f"{title_prefix}：五模型资源占用（按本次硬件口径）", fontsize=14)
    paths["resource_usage.png"] = destination / "resource_usage.png"
    _save_figure(figure, paths["resource_usage.png"], dpi)
    return paths


def _safe_sheet_frame(frame: pd.DataFrame, row_limit: int) -> pd.DataFrame:
    result = frame.head(row_limit).copy()
    for column in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[column]):
            result[column] = result[column].dt.tz_localize(None)
    return result


def write_backtest_workbook(
    *,
    path: Path,
    comparison: pd.DataFrame,
    rank_ic_summary: pd.DataFrame,
    portfolio: MLPortfolioResult,
    factor_correlation_pairs: pd.DataFrame,
    label_coverage_summary: pd.DataFrame,
    label_coverage_by_date: pd.DataFrame,
    report_resource_reference: pd.DataFrame,
    report_performance_reference: pd.DataFrame,
    data_audit_tables: Mapping[str, pd.DataFrame],
    chart_data: Mapping[str, pd.DataFrame],
    chart_paths: Mapping[str, Path],
    row_limit: int,
) -> None:
    """生成多sheet Excel，并将图和对应数据放在同一sheet。"""
    sheets: dict[str, pd.DataFrame] = {
        "核心对照": comparison,
        "研报表现基准": report_performance_reference,
        "研报资源基准": report_resource_reference,
        "RankIC原始": _safe_sheet_frame(portfolio.rank_ic_path_series, row_limit),
        "RankIC摘要": rank_ic_summary,
        "十分档原始": portfolio.path_metric_aggregate.loc[
            portfolio.path_metric_aggregate["period"].eq("all")
            & portfolio.path_metric_aggregate["portfolio"].str.startswith("group_")
        ],
        "单调性": portfolio.monotonicity,
        "预测原始": _safe_sheet_frame(portfolio.factor_panel, row_limit),
        "相关性模型对": factor_correlation_pairs,
        "标签覆盖摘要": label_coverage_summary,
        "标签覆盖逐日": _safe_sheet_frame(label_coverage_by_date, row_limit),
        "年度原始": portfolio.path_metric_aggregate.loc[
            ~portfolio.path_metric_aggregate["period"].eq("all")
        ],
    }
    chart_sheet_map = {
        "feature_importance": "特征重要性图数据",
        "ic_series": "RankIC图数据",
        "group_cumulative_returns": "十分档图数据",
        "train_val_loss": "训练损失图数据",
        "prediction_scatter": "预测散点图数据",
        "model_correlation": "模型相关性图数据",
        "annual_returns": "年度表现图数据",
        "resource_usage": "资源占用图数据",
    }
    sheets.update({chart_sheet_map[key]: frame for key, frame in chart_data.items()})
    for name, frame in data_audit_tables.items():
        sheets[f"数据_{name}"[:31]] = frame
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            if sheet_name in chart_sheet_map.values():
                if len(frame) > row_limit:
                    raise ValueError(f"图数据sheet {sheet_name} 超出展示上限，禁止静默截断")
                excel_frame = frame.copy()
            else:
                excel_frame = _safe_sheet_frame(frame, row_limit)
            _safe_sheet_frame(excel_frame, len(excel_frame)).to_excel(
                writer, sheet_name=sheet_name, index=False
            )

    workbook = load_workbook(destination)
    image_map = {
        "资源占用图数据": "resource_usage.png",
        "RankIC图数据": "ic_series.png",
        "十分档图数据": "group_cumulative_returns.png",
        "训练损失图数据": "train_val_loss.png",
        "特征重要性图数据": "feature_importance.png",
        "预测散点图数据": "prediction_scatter.png",
        "模型相关性图数据": "model_correlation.png",
        "年度表现图数据": "annual_returns.png",
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for index, column_cells in enumerate(sheet.columns, start=1):
            values = [str(cell.value) for cell in list(column_cells)[:200] if cell.value is not None]
            width = min(max([len(value) for value in values] + [8]) + 2, 40)
            sheet.column_dimensions[get_column_letter(index)].width = width
        chart_name = image_map.get(sheet.title)
        if chart_name is not None:
            chart = ExcelImage(chart_paths[chart_name])
            chart.width = 720
            chart.height = 360
            sheet.add_image(chart, "H2")
    workbook.save(destination)


def write_detail_tables(
    *,
    output_dir: Path,
    predictions: pd.DataFrame,
    portfolio: MLPortfolioResult,
    factor_correlation: pd.DataFrame,
    factor_correlation_pairs: pd.DataFrame,
    factor_correlation_by_date: pd.DataFrame,
    label_coverage_summary: pd.DataFrame,
    label_coverage_by_date: pd.DataFrame,
    chart_data: Mapping[str, pd.DataFrame],
) -> dict[str, Path]:
    """完整明细以 Parquet 落盘，不受Excel展示行数限制。"""
    destination = Path(output_dir)
    tables = {
        "predictions.parquet": predictions,
        "factor_panel.parquet": portfolio.factor_panel,
        "factor_correlation.parquet": factor_correlation,
        "factor_correlation_pairs.parquet": factor_correlation_pairs,
        "factor_correlation_by_date.parquet": factor_correlation_by_date,
        "factor_correlation_by_date_m4_legacy.parquet": portfolio.factor_correlation_by_date,
        "label_coverage_summary.parquet": label_coverage_summary,
        "label_coverage_by_date.parquet": label_coverage_by_date,
        "rank_ic.parquet": portfolio.rank_ic,
        "rank_ic_path_series.parquet": portfolio.rank_ic_path_series,
        "rank_ic_path_summary.parquet": portfolio.rank_ic_path_summary,
        "rank_ic_path_aggregate.parquet": portfolio.rank_ic_path_aggregate,
        "group_assignments.parquet": portfolio.group_assignments,
        "path_holdings.parquet": portfolio.path_holdings,
        "path_trades.parquet": portfolio.path_trades,
        "path_nav.parquet": portfolio.path_nav,
        "portfolio_returns.parquet": portfolio.portfolio_returns,
        "path_metrics.parquet": portfolio.path_metrics,
        "annual_metrics.parquet": portfolio.annual_metrics,
        "path_metric_aggregate.parquet": portfolio.path_metric_aggregate,
        "path_daily_aggregate.parquet": portfolio.path_daily_aggregate,
        "monotonicity.parquet": portfolio.monotonicity,
    }
    paths: dict[str, Path] = {}
    for filename, frame in tables.items():
        path = destination / filename
        frame.to_parquet(path, index=False)
        paths[filename] = path
    chart_dir = destination / "chart_data"
    chart_dir.mkdir(parents=True, exist_ok=True)
    for chart_name, frame in chart_data.items():
        path = chart_dir / f"{chart_name}.parquet"
        frame.to_parquet(path, index=False)
        paths[f"chart_data/{chart_name}.parquet"] = path
    return paths
